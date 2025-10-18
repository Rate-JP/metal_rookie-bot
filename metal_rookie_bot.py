# metal_rookie_bot.py  — Excel(.xlsx) 保存版
import os
import asyncio
import logging
from datetime import datetime, timedelta, timezone
from typing import Tuple

import discord
from discord.ext import commands
from dotenv import load_dotenv

# Excel(.xlsx)
try:
    from openpyxl import Workbook, load_workbook
except ImportError as e:
    raise SystemExit("openpyxl が必要です。`pip install openpyxl` を実行してください。") from e

# =====================
# .env
# =====================
load_dotenv()
DISCORD_TOKEN = os.getenv("DISCORD_TOKEN")
CHANNEL_ID = int(os.getenv("CHANNEL_ID", "0"))

# =====================
# 定数・設定
# =====================
PREFIX = "!"
JST = timezone(timedelta(hours=9))
START_ANCHOR = datetime(2025, 10, 16, 12, 0, 0, tzinfo=JST)  # アンカーはJST
INTERVAL = timedelta(hours=2, minutes=30)

MESSAGE_MAIN = "🪙 メタルーキーの時間です！"
EXCEL_PATH = "metal_rookie_bot.xlsx"  # 自動作成
SHEET_NAME = "settings"

# =====================
# ログ設定
# =====================
logging.basicConfig(level=logging.INFO, format="[%(asctime)s] %(levelname)s: %(message)s")
logger = logging.getLogger("metal-rookie-bot")

# =====================
# ユーティリティ
# =====================
def to_jst(dt: datetime) -> datetime:
    """
    任意の datetime を JST に変換。
    - tzなし(naive)は「UTCの値」とみなしてからJSTへ変換（コンテナがUTCでもズレない）
    - tzあり(aware)はそのTZからJSTへ変換
    """
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(JST)


def now_jst() -> datetime:
    """UTCの現在時刻をJSTへ変換（システムローカルTZに依存しない）。"""
    return datetime.now(timezone.utc).astimezone(JST)


def normalize_anchor(anchor: datetime) -> datetime:
    """アンカーをJSTへ正規化（既にJSTでも安全にそのまま返る）。"""
    return to_jst(anchor)


def next_boundary_after(now_jst_val: datetime, anchor_jst: datetime, interval: timedelta) -> datetime:
    """アンカー基準の次の境界（anchor + n*interval、境界上なら now_jst_val）。"""
    if now_jst_val <= anchor_jst:
        return anchor_jst
    elapsed = now_jst_val - anchor_jst
    remainder = elapsed - (elapsed // interval) * interval
    return now_jst_val if remainder == timedelta(0) else now_jst_val + (interval - remainder)


def compute_next_event(
    now_jst_: datetime, anchor_jst: datetime, interval: timedelta, lead_min: int
) -> Tuple[datetime, str, datetime, int]:
    """
    次イベント（'pre' or 'main'）と時刻を返す。
    戻り値: (next_time, kind, boundary, lead_used)
    """
    boundary = next_boundary_after(now_jst_, anchor_jst, interval)
    pre_time = boundary - timedelta(minutes=lead_min)

    if now_jst_ < pre_time:
        return pre_time, "pre", boundary, lead_min
    if now_jst_ == pre_time:
        return now_jst_, "pre", boundary, lead_min
    if now_jst_ < boundary:
        return boundary, "main", boundary, lead_min

    # 境界を過ぎていたら次の境界で再計算
    next_b = next_boundary_after(now_jst_, anchor_jst, interval)
    pre2 = next_b - timedelta(minutes=lead_min)
    if now_jst_ <= pre2:
        return pre2, "pre", next_b, lead_min
    return next_b, "main", next_b, lead_min


def human_delta(td: timedelta) -> str:
    """timedelta を「n時間m分s秒」の日本語に整形。"""
    secs = int(td.total_seconds())
    if secs < 0:
        secs = 0
    h, r = divmod(secs, 3600)
    m, s = divmod(r, 60)
    if h > 0:
        return f"{h}時間{m}分{s}秒"
    if m > 0:
        return f"{m}分{s}秒"
    return f"{s}秒"

# =====================
# Excel(.xlsx): 設定の永続化（3〜15分の事前通知）
#  - .xlsx は内部XMLがUTF-8。書込時は UTF-8 でエンコード可能か検証してから保存します。
# =====================
class SettingsStore:
    def __init__(self, xlsx_path: str, sheet_name: str = SHEET_NAME):
        self.xlsx_path = xlsx_path
        self.sheet_name = sheet_name

    @staticmethod
    def _utf8(s: str) -> str:
        """UTF-8でエンコード可能か検証（明示的UTF-8対応）。"""
        if s is None:
            s = ""
        if not isinstance(s, str):
            s = str(s)
        # ここで例外が出ればUTF-8不可のため上位で扱う
        s.encode("utf-8")
        return s

    def ensure(self) -> None:
        """Excelファイル/シートと初期行を用意。"""
        if not os.path.exists(self.xlsx_path):
            wb = Workbook()
            ws = wb.active
            ws.title = self.sheet_name
            ws.append(["id", "lead_minutes", "updated_at", "encoding"])
            ws.append([1, 5, self._utf8(now_jst().strftime('%Y-%m-%d %H:%M:%S JST')), "UTF-8"])
            wb.save(self.xlsx_path)
            return

        wb = load_workbook(self.xlsx_path)
        if self.sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(self.sheet_name)
            ws.append(["id", "lead_minutes", "updated_at", "encoding"])
            ws.append([1, 5, self._utf8(now_jst().strftime('%Y-%m-%d %H:%M:%S JST')), "UTF-8"])
            wb.save(self.xlsx_path)
            return

        ws = wb[self.sheet_name]
        # ヘッダが無い/壊れている場合に補修
        if ws.max_row == 0:
            ws.append(["id", "lead_minutes", "updated_at", "encoding"])
        # id=1 の行がなければ作る
        has_row = False
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            if row and row[0] == 1:
                has_row = True
                break
        if not has_row:
            ws.append([1, 5, self._utf8(now_jst().strftime('%Y-%m-%d %H:%M:%S JST')), "UTF-8"])
        wb.save(self.xlsx_path)

    def get_lead_minutes(self) -> int:
        """現在の事前通知分を取得（無ければ5）。"""
        try:
            wb = load_workbook(self.xlsx_path, data_only=True)
            ws = wb[self.sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0] == 1:
                    val = int(row[1]) if row[1] is not None else 5
                    return max(3, min(15, val))  # 異常値をはじく
        except Exception as e:
            logger.warning(f"Excel読み込みに失敗しました。既定値(5)を返します: {e}")
        return 5

    def set_lead_minutes(self, minutes: int) -> None:
        """事前通知分を設定（3〜15）。UTF-8文字列で時刻を書込。"""
        if not (3 <= minutes <= 15):
            raise ValueError("lead_minutes は 3〜15 の範囲で指定してください")
        self.ensure()
        wb = load_workbook(self.xlsx_path)
        ws = wb[self.sheet_name]

        # id=1 の行を探して更新
        target_row = None
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == 1:
                target_row = r
                break
        if target_row is None:
            target_row = ws.max_row + 1
            ws.cell(row=target_row, column=1, value=1)

        ws.cell(row=target_row, column=2, value=minutes)
        ts = self._utf8(now_jst().strftime('%Y-%m-%d %H:%M:%S JST'))
        ws.cell(row=target_row, column=3, value=ts)
        ws.cell(row=target_row, column=4, value="UTF-8")
        wb.save(self.xlsx_path)

store = SettingsStore(EXCEL_PATH)
CONFIG_UPDATED = asyncio.Event()  # 設定変更をスケジューラへ即時反映

# =====================
# Discord クライアント（!コマンド）
# =====================
def make_intents() -> discord.Intents:
    intents = discord.Intents.default()
    intents.message_content = True  # これがないと !コマンド検知できません
    intents.messages = True         # 明示
    intents.guilds = True           # 明示
    return intents

bot = commands.Bot(command_prefix=PREFIX, intents=make_intents(), help_command=None)

async def ensure_channel(client: discord.Client, channel_id: int) -> discord.abc.Messageable:
    ch = client.get_channel(channel_id)
    if ch is None:
        ch = await client.fetch_channel(channel_id)
    return ch

async def safe_send(channel: discord.abc.Messageable, content: str) -> None:
    """送信＋例外処理（ログは日本語）。"""
    try:
        # 送信テキストは明示的にUTF-8検証（念のため）
        content.encode("utf-8")
        await channel.send(content)
        logger.info("メッセージを送信しました。")
    except Exception as e:
        logger.exception(f"メッセージの送信に失敗しました: {e}")

def build_help_text(lead_minutes: int) -> str:
    return "\n".join(
        [
            "**📣 BOT起動: 利用できる「!」コマンド**",
            f"現在の事前通知: **{lead_minutes} 分前**",
            "",
            f"• `{PREFIX}notice_get` — 現在の事前通知（分前）を表示",
            f"• `{PREFIX}notice_set <3-15>` — 事前通知の分数を設定（3〜15 以外はエラー）",
            f"• `{PREFIX}next` — 次に発生する 事前通知/本通知 の JST 時刻と残り時間を表示",
            f"• `{PREFIX}help` — このヘルプを表示",
        ]
    )

# =====================
# スケジューラ
# =====================
async def scheduler() -> None:
    await bot.wait_until_ready()
    store.ensure()
    anchor = normalize_anchor(START_ANCHOR)
    channel = await ensure_channel(bot, CHANNEL_ID)

    logger.info(
        f"スケジューラを開始しました (anchor={anchor.isoformat()}, interval={INTERVAL})."
    )

    while not bot.is_closed():
        now = now_jst()
        lead = store.get_lead_minutes()
        next_time, kind, boundary, lead_used = compute_next_event(now, anchor, INTERVAL, lead)

        # 次の通知予定をログ出力（JST基準）
        logger.info(
            "次の通知時刻(JST): %s / 種別=%s / 事前=%s分前",
            next_time.strftime('%Y-%m-%d %H:%M:%S'),
            '事前通知' if kind == 'pre' else '本通知',
            lead_used,
        )

        sleep_sec = max(0.0, (next_time - now).total_seconds())
        logger.info("スリープ: %.1f 秒", sleep_sec)

        # 設定変更を即時反映（擬似スリープ）
        try:
            await asyncio.wait_for(CONFIG_UPDATED.wait(), timeout=sleep_sec)
            CONFIG_UPDATED.clear()
            continue  # ループ先頭へ戻って再計算
        except asyncio.TimeoutError:
            pass

        # 送信
        if kind == "pre":
            await safe_send(channel, f"🌈 メタルーキーまであと{lead_used}分です！")
        else:
            await safe_send(channel, MESSAGE_MAIN)

        # 送信直後に次回予定をログ
        now_after = now_jst()
        lead_after = store.get_lead_minutes()
        next_time_after, kind_after, _, _ = compute_next_event(
            now_after, anchor, INTERVAL, lead_after
        )
        logger.info(
            "次回の通知(JST): %s / 種別=%s / 事前=%s分前",
            next_time_after.strftime('%Y-%m-%d %H:%M:%S'),
            '事前通知' if kind_after == 'pre' else '本通知',
            lead_after,
        )

# =====================
# コマンド
# =====================
@bot.command(name="notice_get")
async def notice_get(ctx: commands.Context) -> None:
    """現在の事前通知（分前）を表示"""
    store.ensure()
    m = store.get_lead_minutes()
    await ctx.reply(f"ℹ️ 現在の事前通知は **{m} 分前**です。", mention_author=False)

@bot.command(name="notice_set")
async def notice_set_cmd(ctx: commands.Context, minutes: int | None = None) -> None:
    """
    事前通知の分数を設定（3〜15） 使い方: !notice_set 10
    """
    store.ensure()
    if minutes is None:
        await ctx.reply(
            f"使い方: `{PREFIX}notice_set <分>` 例: `{PREFIX}notice_set 10`",
            mention_author=False,
        )
        return

    if not (3 <= minutes <= 15):
        await ctx.reply("⚠️ 通知時間は **3〜15分前** でのみ設定できます。", mention_author=False)
        return

    try:
        store.set_lead_minutes(minutes)
        await ctx.reply(f"✅ 事前通知を **{minutes} 分前**に設定しました。", mention_author=False)
        CONFIG_UPDATED.set()  # スケジューラに即時反映
    except Exception as e:
        logger.exception(e)
        await ctx.reply("❌ 設定に失敗しました。ログを確認してください。", mention_author=False)

@bot.command(name="next")
async def next_cmd(ctx: commands.Context) -> None:
    """次の通知（事前/本）と各時刻を表示"""
    store.ensure()
    now = now_jst()
    lead = store.get_lead_minutes()
    anchor = normalize_anchor(START_ANCHOR)

    next_time, kind, boundary, _ = compute_next_event(now, anchor, INTERVAL, lead)
    next_main = boundary
    pre_time = boundary - timedelta(minutes=lead)
    next_pre = pre_time if now < pre_time else boundary + INTERVAL - timedelta(minutes=lead)

    eta_next = human_delta(next_time - now)
    eta_pre = human_delta(next_pre - now)
    eta_main = human_delta(next_main - now)

    text = "\n".join(
        [
            f"🗓 現在の設定: 事前通知 **{lead} 分前**",
            f"⏳ 次の事前通知: {next_pre.strftime('%Y-%m-%d %H:%M:%S')} JST（あと {eta_pre}）",
            f"⏰ 次の本通知:   {next_main.strftime('%Y-%m-%d %H:%M:%S')} JST（あと {eta_main}）",
        ]
    )
    await ctx.reply(text, mention_author=False)

@bot.command(name="help")
async def help_cmd(ctx: commands.Context) -> None:
    """コマンド一覧を表示"""
    store.ensure()
    lead = store.get_lead_minutes()
    await ctx.reply(build_help_text(lead), mention_author=False)

# =====================
# イベント
# =====================
@bot.event
async def on_ready():
    logger.info(f"ログインに成功: {bot.user} (ID: {bot.user.id})")

    # 起動時ヘルプ送信
    store.ensure()
    lead = store.get_lead_minutes()
    try:
        channel = await ensure_channel(bot, CHANNEL_ID)
        await safe_send(channel, build_help_text(lead))
        logger.info("起動時のヘルプメッセージを送信しました。")
    except Exception as e:
        logger.exception(f"起動時のヘルプメッセージ送信に失敗しました: {e}")

    # スケジューラ起動
    asyncio.create_task(scheduler())

# =====================
# エントリポイント
# =====================
if __name__ == "__main__":
    if not DISCORD_TOKEN or CHANNEL_ID == 0:
        raise SystemExit("環境変数 DISCORD_TOKEN / CHANNEL_ID を設定してください（.env 参照）")

    # 先に Excel を確実に初期化
    store.ensure()

    bot.run(DISCORD_TOKEN)
