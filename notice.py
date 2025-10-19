import os
import asyncio
import logging
import inspect
from datetime import datetime, timedelta, timezone
from typing import Tuple, Optional

import discord
from discord.ext import commands
from dotenv import load_dotenv

# Excel(.xlsx)
try:
    from openpyxl import Workbook, load_workbook
except ImportError as e:
    raise SystemExit("openpyxl が必要です。`pip install openpyxl` を実行してください。") from e

logger = logging.getLogger("metal-rookie-bot")

# ---------------------
# 共通設定（環境変数）
# ---------------------
load_dotenv()
PREFIX = os.getenv("PREFIX", "!")
CHANNEL_ID = int(os.getenv("CHANNEL_ID", "0"))

# Anchors / Interval は元コードと同一値を採用
JST = timezone(timedelta(hours=9))
START_ANCHOR = datetime(2025, 10, 16, 12, 0, 0, tzinfo=JST)  # アンカーはJST（固定）
INTERVAL = timedelta(hours=2, minutes=30)

MESSAGE_MAIN = os.getenv("MESSAGE_MAIN", "🪙 メタルーキーの時間です！")
EXCEL_PATH = os.getenv("EXCEL_PATH", "metal_rookie_bot.xlsx")
SHEET_NAME = os.getenv("SHEET_NAME", "settings")

# 値域・制限
LEAD_MIN = 3
LEAD_MAX = 15
DEFAULT_LEAD = 10
MAX_DISCORD_MESSAGE_LEN = 2000

# ---------------------
# JSTユーティリティ（コンテナTZに依存しない）
# ---------------------
def to_jst(dt: datetime) -> datetime:
    """任意の tz-aware/naive datetime を JST へ変換（naive は UTC 前提で扱う）"""
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(JST)

def now_jst() -> datetime:
    """JST 現在時刻"""
    return datetime.now(timezone.utc).astimezone(JST)

def normalize_anchor(anchor: datetime) -> datetime:
    """アンカーを JST に正規化"""
    return to_jst(anchor)

def next_boundary_after(now_jst_val: datetime, anchor_jst: datetime, interval: timedelta) -> datetime:
    """now 以降で最初の境界（アンカー + n*interval）を返す。境界上なら now を返す。"""
    if now_jst_val <= anchor_jst:
        return anchor_jst
    elapsed = now_jst_val - anchor_jst
    remainder = elapsed - (elapsed // interval) * interval
    return now_jst_val if remainder == timedelta(0) else now_jst_val + (interval - remainder)

def compute_next_event(
    now_jst_: datetime, anchor_jst: datetime, interval: timedelta, lead_min: int
) -> Tuple[datetime, str, datetime, int]:
    """
    Returns:
        (next_time, kind, boundary, lead_used)
        kind in {"pre", "main"}
    """
    boundary = next_boundary_after(now_jst_, anchor_jst, interval)
    pre_time = boundary - timedelta(minutes=lead_min)

    if now_jst_ < pre_time:
        return pre_time, "pre", boundary, lead_min
    if now_jst_ == pre_time:
        return now_jst_, "pre", boundary, lead_min
    if now_jst_ < boundary:
        return boundary, "main", boundary, lead_min

    next_b = next_boundary_after(now_jst_, anchor_jst, interval)
    pre2 = next_b - timedelta(minutes=lead_min)
    if now_jst_ <= pre2:
        return pre2, "pre", next_b, lead_min
    return next_b, "main", next_b, lead_min

def human_delta(td: timedelta) -> str:
    secs = int(td.total_seconds())
    if secs < 0:
        secs = 0
    h, r = divmod(secs, 3600)
    m, s = divmod(r, 60)
    if h > 0:
        return f"{h}時間{m}分{s}秒"
    if m > 0:
        return f"{m}分{s}秒"
    return f"{s}秒"

# ---------------------
# 送信ユーティリティ
# ---------------------
def _chunk_message(content: str, limit: int = MAX_DISCORD_MESSAGE_LEN):
    """Discord の 2000 文字制限に合わせて分割"""
    content = content or ""
    for i in range(0, len(content), limit):
        yield content[i : i + limit]

async def ensure_channel(client: discord.Client, channel_id: int) -> discord.abc.Messageable:
    ch = client.get_channel(channel_id)
    if ch is None:
        ch = await client.fetch_channel(channel_id)
    return ch

async def safe_send(channel: discord.abc.Messageable, content: str) -> None:
    try:
        content.encode("utf-8")
        parts = list(_chunk_message(content))
        for idx, p in enumerate(parts, 1):
            await channel.send(p)
            if len(parts) > 1:
                logger.info("メッセージを分割送信しました (%s/%s)", idx, len(parts))
        logger.info("メッセージを送信しました。")
    except Exception as e:
        logger.exception(f"メッセージ送信に失敗しました: {e}")

def build_help_text(lead_minutes: int) -> str:
    return "\n".join(
        [
            "**📣 メタルーキーお知らせ機能**",
            f"• 現在の事前通知: **{lead_minutes} 分前**",
            f"• 設定ファイル: `{EXCEL_PATH}` / シート: `{SHEET_NAME}`",
            "",
            f"• `{PREFIX}notice_get` — 現在の事前通知（分前）を表示",
            f"• `{PREFIX}notice_set <{LEAD_MIN}-{LEAD_MAX}>` — 事前通知の分数を設定",
            f"• `{PREFIX}next` — 次に発生する 事前通知/本通知 の JST 時刻と残り時間を表示",
            f"• `{PREFIX}help` — このヘルプを表示",
        ]
    )

# ---------------------
# Excel 設定ストア
# ---------------------
class SettingsStore:
    def __init__(self, xlsx_path: str, sheet_name: str = SHEET_NAME):
        self.xlsx_path = xlsx_path
        self.sheet_name = sheet_name

    @staticmethod
    def _utf8(s: Optional[str]) -> str:
        if s is None:
            s = ""
        if not isinstance(s, str):
            s = str(s)
        s.encode("utf-8")
        return s

    def _ensure_dir(self) -> None:
        d = os.path.dirname(os.path.abspath(self.xlsx_path))
        if d and not os.path.exists(d):
            os.makedirs(d, exist_ok=True)

    def ensure(self) -> None:
        """設定ファイル・シート・ヘッダ・既定行を保証"""
        self._ensure_dir()

        if not os.path.exists(self.xlsx_path):
            wb = Workbook()
            ws = wb.active
            ws.title = self.sheet_name
            ws.append(["id", "lead_minutes", "updated_at", "encoding"])
            ws.append([1, DEFAULT_LEAD, self._utf8(now_jst().strftime('%Y-%m-%d %H:%M:%S JST')), "UTF-8"])
            wb.save(self.xlsx_path)
            wb.close()
            return

        wb = load_workbook(self.xlsx_path)
        try:
            if self.sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(self.sheet_name)
                ws.append(["id", "lead_minutes", "updated_at", "encoding"])
                ws.append([1, DEFAULT_LEAD, self._utf8(now_jst().strftime('%Y-%m-%d %H:%M:%S JST')), "UTF-8"])
                wb.save(self.xlsx_path)
                return

            ws = wb[self.sheet_name]
            # ヘッダ保証
            header_ok = (
                ws.cell(row=1, column=1).value == "id"
                and ws.cell(row=1, column=2).value == "lead_minutes"
                and ws.cell(row=1, column=3).value == "updated_at"
                and ws.cell(row=1, column=4).value == "encoding"
            )
            if not header_ok:
                ws.delete_rows(1, ws.max_row)  # 破損時は作り直し
                ws.append(["id", "lead_minutes", "updated_at", "encoding"])

            # 既定行保証（id=1）
            has_row = False
            for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
                if row and row[0] == 1:
                    has_row = True
                    break
            if not has_row:
                ws.append([1, DEFAULT_LEAD, self._utf8(now_jst().strftime('%Y-%m-%d %H:%M:%S JST')), "UTF-8"])

            wb.save(self.xlsx_path)
        finally:
            wb.close()

    def get_lead_minutes(self) -> int:
        try:
            wb = load_workbook(self.xlsx_path, data_only=True)
            try:
                ws = wb[self.sheet_name]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and row[0] == 1:
                        val = int(row[1]) if row[1] is not None else DEFAULT_LEAD
                        return max(LEAD_MIN, min(LEAD_MAX, val))
            finally:
                wb.close()
        except Exception as e:
            logger.warning(f"Excel読み込み失敗。既定値({DEFAULT_LEAD})を返します: {e}")
        return DEFAULT_LEAD

    def set_lead_minutes(self, minutes: int) -> None:
        if not (LEAD_MIN <= minutes <= LEAD_MAX):
            raise ValueError(f"lead_minutes は {LEAD_MIN}〜{LEAD_MAX} の範囲で指定してください")
        self.ensure()
        wb = load_workbook(self.xlsx_path)
        try:
            ws = wb[self.sheet_name]

            target_row = None
            for r in range(2, ws.max_row + 1):
                if ws.cell(row=r, column=1).value == 1:
                    target_row = r
                    break
            if target_row is None:
                target_row = ws.max_row + 1
                ws.cell(row=target_row, column=1, value=1)

            ws.cell(row=target_row, column=2, value=minutes)
            ts = self._utf8(now_jst().strftime('%Y-%m-%d %H:%M:%S JST'))
            ws.cell(row=target_row, column=3, value=ts)
            ws.cell(row=target_row, column=4, value="UTF-8")
            wb.save(self.xlsx_path)
        finally:
            wb.close()

    # ---- 非同期（to_thread で I/O をイベントループから隔離）
    async def ensure_async(self) -> None:
        await asyncio.to_thread(self.ensure)

    async def get_lead_minutes_async(self) -> int:
        return await asyncio.to_thread(self.get_lead_minutes)

    async def set_lead_minutes_async(self, minutes: int) -> None:
        await asyncio.to_thread(self.set_lead_minutes, minutes)

# ---------------------
# Cog: 通知機能
# ---------------------
class MetalRookieCog(commands.Cog):
    def __init__(self, bot: commands.Bot):
        self.bot = bot
        self.store = SettingsStore(EXCEL_PATH)
        self.anchor = normalize_anchor(START_ANCHOR)
        self.interval = INTERVAL
        self.message_main = MESSAGE_MAIN
        self.channel_id = CHANNEL_ID
        self._scheduler_task: Optional[asyncio.Task] = None
        self._ready_once = False
        self.CONFIG_UPDATED = asyncio.Event()

    def _attach_task_logger(self, task: asyncio.Task) -> None:
        def _done_cb(t: asyncio.Task):
            try:
                _ = t.result()
            except asyncio.CancelledError:
                logger.info("スケジューラタスクはキャンセルされました。")
            except Exception:
                logger.exception("スケジューラタスクで未処理例外が発生しました。")
        task.add_done_callback(_done_cb)

    # ---- Bot ready → 起動時に一度だけスケジューラ開始 & ヘルプ送付
    @commands.Cog.listener()
    async def on_ready(self):
        if self._ready_once:
            return
        self._ready_once = True

        if self.channel_id == 0:
            logger.error("環境変数 CHANNEL_ID が未設定です。送信できません。")
            return

        # Excel 初期化（非同期）
        await self.store.ensure_async()

        # 起動時ヘルプ
        try:
            ch = await ensure_channel(self.bot, self.channel_id)
            lead = await self.store.get_lead_minutes_async()
            await safe_send(ch, build_help_text(lead))
            logger.info("起動時ヘルプを送信しました。")
        except Exception as e:
            logger.exception(f"起動時ヘルプ送信に失敗しました: {e}")

        # スケジューラ起動
        self._scheduler_task = asyncio.create_task(self._scheduler_loop(), name="metal_rookie_scheduler")
        self._attach_task_logger(self._scheduler_task)

    # ---- Cog アンロード時にタスクを確実に停止
    def cog_unload(self):
        if self._scheduler_task and not self._scheduler_task.done():
            self._scheduler_task.cancel()

    # ---- スケジューラ本体
    async def _scheduler_loop(self) -> None:
        try:
            ch = await ensure_channel(self.bot, self.channel_id)
            logger.info(
                "スケジューラ開始 (anchor=%s, interval=%s).",
                self.anchor.isoformat(),
                self.interval,
            )
            while not self.bot.is_closed():
                now = now_jst()
                lead = await self.store.get_lead_minutes_async()
                next_time, kind, boundary, lead_used = compute_next_event(now, self.anchor, self.interval, lead)

                logger.info(
                    "次の通知(JST): %s / 種別=%s / 事前=%s分前",
                    next_time.strftime('%Y-%m-%d %H:%M:%S'),
                    '事前通知' if kind == 'pre' else '本通知',
                    lead_used,
                )

                sleep_sec = max(0.0, (next_time - now).total_seconds())
                logger.info("スリープ: %.1f 秒", sleep_sec)

                try:
                    await asyncio.wait_for(self.CONFIG_UPDATED.wait(), timeout=sleep_sec)
                    self.CONFIG_UPDATED.clear()
                    # 設定変更が入ったので次ループで再計算
                    continue
                except asyncio.TimeoutError:
                    pass

                # 送信
                if kind == "pre":
                    await safe_send(ch, f"🌈 メタルーキーまであと{lead_used}分です！")
                else:
                    await safe_send(ch, self.message_main)

                # 次回の予定をログ
                now_after = now_jst()
                lead_after = await self.store.get_lead_minutes_async()
                next_time_after, kind_after, _, _ = compute_next_event(
                    now_after, self.anchor, self.interval, lead_after
                )
                logger.info(
                    "次回(JST): %s / 種別=%s / 事前=%s分前",
                    next_time_after.strftime('%Y-%m-%d %H:%M:%S'),
                    '事前通知' if kind_after == 'pre' else '本通知',
                    lead_after,
                )
        except asyncio.CancelledError:
            logger.info("スケジューラループがキャンセルされました。")
            raise
        except Exception:
            logger.exception("スケジューラループで例外が発生しました。")

    # ---- コマンド群
    @commands.command(name="notice_get")
    async def notice_get(self, ctx: commands.Context) -> None:
        await self.store.ensure_async()
        m = await self.store.get_lead_minutes_async()
        await ctx.reply(f"ℹ️ 現在の事前通知は **{m} 分前**です。", mention_author=False)

    @commands.command(name="notice_set")
    async def notice_set_cmd(self, ctx: commands.Context, minutes: Optional[int] = None) -> None:
        await self.store.ensure_async()
        if minutes is None:
            await ctx.reply(
                f"使い方: `{PREFIX}notice_set <分>` 例: `{PREFIX}notice_set 10`  （{LEAD_MIN}〜{LEAD_MAX} のみ）",
                mention_author=False,
            )
            return

        if not (LEAD_MIN <= minutes <= LEAD_MAX):
            await ctx.reply(f"⚠️ 通知時間は **{LEAD_MIN}〜{LEAD_MAX}分前** でのみ設定できます。", mention_author=False)
            return

        try:
            await self.store.set_lead_minutes_async(minutes)
            await ctx.reply(f"✅ 事前通知を **{minutes} 分前**に設定しました。", mention_author=False)
            self.CONFIG_UPDATED.set()
        except Exception as e:
            logger.exception(e)
            await ctx.reply("❌ 設定に失敗しました。ログを確認してください。", mention_author=False)

    @commands.command(name="next")
    async def next_cmd(self, ctx: commands.Context) -> None:
        await self.store.ensure_async()
        now = now_jst()
        lead = await self.store.get_lead_minutes_async()
        anchor = self.anchor

        next_time, kind, boundary, _ = compute_next_event(now, anchor, self.interval, lead)
        next_main = boundary
        pre_time = boundary - timedelta(minutes=lead)
        next_pre = pre_time if now < pre_time else boundary + self.interval - timedelta(minutes=lead)

        eta_pre = human_delta(next_pre - now)
        eta_main = human_delta(next_main - now)

        text = "\n".join(
            [
                f"🗓 現在の設定: 事前通知 **{lead} 分前**",
                f"⏳ 次の事前通知: {next_pre.strftime('%Y-%m-%d %H:%M:%S')} JST（あと {eta_pre}）",
                f"⏰ 次の本通知:   {next_main.strftime('%Y-%m-%d %H:%M:%S')} JST（あと {eta_main}）",
            ]
        )
        await ctx.reply(text, mention_author=False)

    @commands.command(name="help")
    async def help_cmd(self, ctx: commands.Context) -> None:
        await self.store.ensure_async()
        lead = await self.store.get_lead_minutes_async()
        await ctx.reply(build_help_text(lead), mention_author=False)

# 拡張エントリ（discord.py のバージョン差異を吸収）
async def setup(bot: commands.Bot):
    cog = MetalRookieCog(bot)
    add_cog = getattr(bot, "add_cog", None)
    if add_cog is None:
        raise RuntimeError("bot.add_cog が見つかりません。discord.py のバージョンを確認してください。")

    if inspect.iscoroutinefunction(add_cog):
        # add_cog がコルーチンの環境
        await add_cog(cog)  # type: ignore[misc]
    else:
        # add_cog が同期関数の環境
        add_cog(cog)  # type: ignore[misc]
